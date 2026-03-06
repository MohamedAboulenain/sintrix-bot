"""
Excel service:
  - extract_text()    — read Excel bytes → plain text (openpyxl)
  - generate_excel()  — structured content → Excel bytes (openpyxl)
"""
from __future__ import annotations

import io
import json
import re


def extract_text(data: bytes) -> tuple[str, int]:
    """Return (full_text, sheet_count) from Excel bytes."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"Sheet: {sheet.title}\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts), len(wb.sheetnames)


def generate_excel(title: str, content: str) -> bytes:
    """
    Convert AI-generated content to a styled Excel workbook.
    Expects content as JSON or markdown table; falls back to plain text rows.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY = "0A0E27"
    CYAN = "22D3EE"
    LIGHT_ROW = "EBF8FF"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # sheet name limit

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=NAVY)
    title_font = Font(bold=True, color=NAVY, size=14)
    alt_fill = PatternFill("solid", fgColor=LIGHT_ROW)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Try to parse as JSON array of objects
    rows_data = _parse_content(content)

    if not rows_data:
        # Fallback: split by newline
        rows_data = [{"Content": line} for line in content.splitlines() if line.strip()]

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(rows_data[0]) if rows_data else 1)}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # Header row
    headers = list(rows_data[0].keys()) if rows_data else ["Content"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 20

    # Data rows
    for row_idx, row_dict in enumerate(rows_data, start=3):
        fill = alt_fill if (row_idx % 2 == 1) else None
        for col, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col, value=str(row_dict.get(key, "")))
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 18

    # Auto-fit column widths
    for col in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, len(rows_data) + 3)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 50)

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_content(content: str) -> list[dict]:
    """Try JSON → markdown table → plain numbered lines."""
    # Strip markdown code fences (```json ... ``` or ``` ... ```)
    stripped = re.sub(r"^```[^\n]*\n?", "", content.strip(), flags=re.MULTILINE)
    stripped = re.sub(r"```$", "", stripped.strip(), flags=re.MULTILINE).strip()

    # JSON array
    for candidate in (stripped, content):
        try:
            data = json.loads(candidate)
            if isinstance(data, list) and data and isinstance(data[0], dict):
                return data
        except Exception:
            pass

    # Markdown table: | col | col | …
    lines = [l.strip() for l in content.splitlines() if l.strip()]
    table_lines = [l for l in lines if l.startswith("|")]
    if len(table_lines) >= 2:
        headers_raw = [h.strip() for h in table_lines[0].strip("|").split("|")]
        rows = []
        for line in table_lines[2:]:  # skip separator
            cells = [c.strip() for c in line.strip("|").split("|")]
            rows.append(dict(zip(headers_raw, cells)))
        if rows:
            return rows

    return []
